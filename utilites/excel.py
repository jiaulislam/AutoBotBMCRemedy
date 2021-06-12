import os
from typing import Dict

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from rich.prompt import Confirm

from utilites import make_data


class OpenExcel:
    """ Safe open excel file """

    def __init__(self, file: str, read_only: bool = False) -> None:
        self._file_path: str = file
        self._read_only: bool = read_only

        if not os.path.exists(file):
            raise FileNotFoundError("File doesn't exists.")

        if not self._read_only:
            while True:
                try:
                    with open(file, "a+"):
                        break
                except IOError:  # File is open, Handle the situation
                    if Confirm.ask("File is in used mode. Is closed ?", default="(y)", show_default=True):
                        break
                    pass

        self._file: Workbook = load_workbook(self._file_path, self._read_only)
        self._sheet: Worksheet = self._file["Main"]

    def __enter__(self) -> Worksheet:
        return self._sheet

    def __exit__(self, exc_type, exc_value, exc_traceback) -> None:
        if not self._read_only:
            self._file.save(self._file_path)
        if exc_type:
            print(exc_type, exc_value, exc_traceback)
        self._file.close()


def get_last_empty_cell(sheet: Worksheet) -> int:
    """ Return the last empty row number in a worksheet by sheet"""
    _last_blank_row: int = 0

    for row in range(1, sheet.max_row):
        if sheet[f"A{row}"].value is None:
            break
        else:
            _last_blank_row += 1
    return _last_blank_row + 1


def get_last_empty_row(file: str) -> int:
    """ Return the last empty row number in a worksheet by file"""
    _last_row: int = 0

    with OpenExcel(file, read_only=True) as sheet:
        for row in range(1, sheet.max_row):
            if sheet[f'A{row}'].value is None:
                break
            _last_row += 1

    return _last_row + 1


def pull(file: str, row_num: int) -> Dict[str, str]:
    """ Pull data from shared excel
        :Args
            file_path: str
                Path to the file
            read_only: bool
                Read only mode for optimized
            row_num: int
                read the spcific number row
        :Returns A key value pair as cell : value
    """
    with OpenExcel(file, read_only=True) as file:
        _cell_data: Dict[str, str] = {}
        for _cell_name in range(65, 76):
            _cell = f"{chr(_cell_name)}{row_num}"
            if _cell[-2] == "F":
                _cell_data[_cell[-2]] = ",".join(make_data.split_string(file[_cell].value))
                continue
            _cell_data[_cell[-2]] = file[_cell].value
        return _cell_data


def put(file: str, data_obj: Dict[str, str]) -> None:
    """ Put the user data to excel in last row
        :Args
            file_path: str
                Path to the file to write
            data_obj: Dict[str, str]
                dictionary object that contains the key value pair data
    """
    with OpenExcel(file) as file:
        _start_row: int = get_last_empty_cell(file)
        for _cell_name in range(65, 76):
            _cell = f"{chr(_cell_name)}{_start_row}"
            file[_cell] = data_obj.get(_cell[-2])
