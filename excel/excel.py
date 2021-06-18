import os
from typing import Dict
from utilites import make_data
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from rich.prompt import Confirm
from excel import IExcel

EXCEL_SHEET_NAME = "MAIN"


class Excel(IExcel):
    def __init__(self, file: str) -> None:
        self._file: str = file

        if not os.path.exists(self._file):
            raise FileNotFoundError("File doesn't exist")

        while True:
            try:
                with open(self._file, "a+"):
                    break
            except IOError:
                if Confirm.ask(
                    "File is in used mode. Is it closed ?",
                    default="(y)",
                    show_default=True,
                ):
                    break
                pass

        self._wb: Workbook = load_workbook(
            filename=self._file, data_only=True
        )
        self._ws: Worksheet = self.__change_sheet(EXCEL_SHEET_NAME)

    def __change_sheet(self, _sheet_name: str) -> Worksheet:
        return self._wb[_sheet_name]

    def get_row(self, row_num: int) -> Dict[str, str]:
        _row_data = {}

        for _cell_chr in range(65, 76):
            _cell_idx = f"{chr(_cell_chr)}{row_num}"
            if chr(_cell_chr) == "F":
                _row_data[chr(_cell_chr)] = ",".join(
                    make_data.split_string(self._ws[_cell_idx].value)
                )
                continue
            _row_data[chr(_cell_chr)] = self._ws[_cell_idx].value

        return _row_data

    def insert(self, row_num: int, cr_number: str) -> None:
        _cell = f"J{row_num}"
        self._ws[_cell] = cr_number

    def get_last_row(self) -> int:
        _last_row = 0

        for row in range(1, self._ws.max_row):
            if self._ws[f"A{row}"].value is None or self._ws[f"A{row}"] == "0":
                break
            _last_row += 1

        return _last_row

    def save(self) -> None:
        self._wb.save(self._file)

    def close(self) -> None:
        self._wb.close()
